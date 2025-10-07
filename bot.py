import os
import shutil
import pandas as pd
import random
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, filters, ContextTypes

TOKEN = "7449004443:AAEVwVSuBiod-6qyAZRszz5livezwTJY21Y"

EXCEL_TEMPLATE = r"C:\Users\User\Desktop\Deutsch\Words code\words.xlsx"
DATA_FOLDER = r"C:\Users\User\Desktop\Deutsch\Words code\user_files"
os.makedirs(DATA_FOLDER, exist_ok=True)

users = {}  # chat_id: {"df":..., "file":..., "last_index":..., "counter":..., "current_word":..., "correct_answer":...}

def create_user_file(chat_id):
    user_file = os.path.join(DATA_FOLDER, f"user_{chat_id}.xlsx")
    if not os.path.exists(user_file):
        # بجای ساخت فایل خالی، کپی فایل اصلی می‌کنیم
        shutil.copy(EXCEL_TEMPLATE, user_file)
    return user_file

def save_progress_inplace(df, excel_file, apply_colors=False):
    wb = load_workbook(excel_file)
    ws = wb.active

    def ensure_col(header):
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=c).value == header:
                return c
        new_c = ws.max_column + 1
        ws.cell(row=1, column=new_c, value=header)
        return new_c

    col_asked = ensure_col("Times Asked")
    col_correct = ensure_col("Times Correct")
    col_success = ensure_col("Success Rate")

    n = len(df)
    for i in range(n):
        ws.cell(row=i+2, column=col_asked, value=df.at[i, "Times Asked"])
        ws.cell(row=i+2, column=col_correct, value=df.at[i, "Times Correct"])
        ws.cell(row=i+2, column=col_success, value=df.at[i, "Success Rate"])

    if apply_colors:
        col_letter = get_column_letter(col_success)
        rng = f"{col_letter}2:{col_letter}{n+1}"
        color_rule = ColorScaleRule(
            start_type='num', start_value=0, start_color='F8696B',
            mid_type='num', mid_value=0.5, mid_color='FFEB84',
            end_type='num', end_value=1, end_color='63BE7B'
        )
        try:
            ws.conditional_formatting._cf_rules.pop(rng, None)
        except Exception:
            pass
        ws.conditional_formatting.add(rng, color_rule)

    wb.save(excel_file)

def pick_word(user_state):
    df = user_state["df"]
    for col in ["Times Asked", "Times Correct", "Success Rate"]:
        if col not in df.columns:
            df[col] = 0 if col != "Success Rate" else 0.0

    df["Success Rate"] = df.apply(
        lambda row: (row["Times Correct"] / row["Times Asked"]) if row["Times Asked"] > 0 else 0.0,
        axis=1
    )
    weights = (1 - df["Success Rate"]).tolist()
    weights = [w if w > 1e-6 else 1e-6 for w in weights]

    index = random.choices(range(len(df)), weights=weights, k=1)[0]
    user_state["last_index"] = index
    user_state["current_word"] = str(df.iloc[index, 0])
    user_state["correct_answer"] = str(df.iloc[index, 1])
    df.at[index, "Times Asked"] += 1
    return user_state["current_word"]

async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.message.chat_id
    user_file = create_user_file(chat_id)
    df = pd.read_excel(user_file)

    users[chat_id] = {
        "df": df,
        "file": user_file,
        "last_index": None,
        "counter": 0,
        "current_word": None,
        "correct_answer": None
    }

    await update.message.reply_text("📚 تمرین لغات شروع شد! برای خروج بنویس exit.")
    word = pick_word(users[chat_id])
    await update.message.reply_text(f"🇩🇪 معنی این کلمه چی میشه\n'{word}'")

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.message.chat_id
    if chat_id not in users:
        await update.message.reply_text("⛔️ لطفاً با /start شروع کن.")
        return

    user_state = users[chat_id]
    df = user_state["df"]
    text = update.message.text.strip().lower()

    if text == "exit":
        save_progress_inplace(df, user_state["file"], apply_colors=True)
        await update.message.reply_text("👋 تمرین تموم شد! موفق باشی 💪")
        del users[chat_id]
        return

    if text == user_state["correct_answer"].lower():
        await update.message.reply_text("✅ آفرین! 👏")
        df.at[user_state["last_index"], "Times Correct"] += 1
    else:
        await update.message.reply_text(f"❌ اشتباه بود ، جواب درست:\n {user_state['correct_answer']}")

    user_state["counter"] += 1
    if user_state["counter"] % 5 == 0:
        save_progress_inplace(df, user_state["file"], apply_colors=False)

    word = pick_word(user_state)
    await update.message.reply_text(f"🇩🇪 معنی این کلمه چی میشه\n'{word}'")

app = ApplicationBuilder().token(TOKEN).build()
app.add_handler(CommandHandler("start", start_command))
app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
app.run_polling()
